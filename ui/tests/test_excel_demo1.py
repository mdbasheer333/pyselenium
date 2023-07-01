import openpyxl


def tsts_excel_reading():
    wb = openpyxl.Workbook()
    wb.create_sheet(index=1, title="demo sheet1")
    ws = wb['demo sheet1']
    ws.cell(1, 1).value = "username"
    ws.cell(1, 2).value = "password"
    ws.cell(2, 1).value = "demo"
    ws.cell(2, 2).value = "demo#1234"
    wb.save("C:\\Users\\RASHEED\\Desktop\\demo_exl.xlsx")


tsts_excel_reading()
