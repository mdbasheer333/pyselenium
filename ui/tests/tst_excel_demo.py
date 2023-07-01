import openpyxl


def tsts_excel_reading():
    wb = openpyxl.load_workbook("C:\\Users\\RASHEED\\Desktop\\sample_exl.xlsx")
    # print(wb.sheetnames)
    # for eacH_sheet_name in wb.sheetnames:
    #     print(eacH_sheet_name)
    ws = wb['samplesheet']
    # print(ws.cell(1, 1).value)
    # print(ws.cell(1, 2).value)
    # print(ws.cell(2, 1).value)
    # print(ws.cell(2, 2).value)
    # print(ws.cell(3, 1).value)
    # print(ws.cell(3, 2).value)

    # print("number of rows are  ", ws.max_row)
    # print("number of cols are  ", ws.max_column)

    total_exl_data = []

    for row in range(2, ws.max_row + 1):
        dct = {}
        for col in range(1, ws.max_column + 1):
            # print(ws.cell(1, col).value)
            # print(ws.cell(row, col).value)
            dct[ws.cell(1, col).value] = ws.cell(row, col).value
        # print(dct)
        total_exl_data.append(dct)

    # print(total_exl_data)

    my_tc_data = {}
    for each_dict in total_exl_data:
        # print(each_dict)
        if each_dict.get("TCID") == "reschedule":
            my_tc_data = each_dict
            break
    print(my_tc_data)


tsts_excel_reading()
